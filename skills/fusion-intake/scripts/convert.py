#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "openpyxl>=3.1.0",
#   "PyYAML>=6.0",
#   "pymupdf>=1.24.0",
#   "pillow>=10.0.0",
# ]
# ///
"""fusion-intake Stage-1 engine: admit / prepare / link / batch / cleanup.

Deterministic, no LLM. `admit` is the ONLY writer of sources/MANIFEST.md in
the whole Fusion reference implementation. `prepare` routes by format:
extractive files (xlsx/csv) become conformant documents directly; everything
else gets a work dir under workbench/.intake/ with page text + images +
manifest.json for the Stage-2 agent (references/convert.md). The page-
coverage invariant — every page recorded, low-text pages flagged
needs_vision — makes silent-empty output structurally impossible. `batch`
runs a validated admit+link op-list in one process for delivery-scale
intake (references/delivery.md) — it reuses `admit`/`link`, never a second
writer.
"""
import argparse
import csv as csv_mod
import hashlib
import json
import posixpath
import re
import shutil
import subprocess
import sys
import uuid
import zipfile
from datetime import datetime
from email import policy
from email.parser import BytesParser
from html.parser import HTMLParser
from pathlib import Path

import yaml

AURORAS = ("commitments", "focus", "ops", "collab", "life", "explore",
           "archive", "library")

EXTRACTIVE_EXTS = {".xlsx", ".csv"}
LIBREOFFICE_EXTS = {".docx", ".pptx", ".doc", ".odt", ".rtf", ".key",
                    ".pages", ".ppt", ".xls", ".html", ".htm"}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
# Tiffs are never stored as tiffs: sources/ would carry needlessly large
# files for no fidelity gain over a lossless PNG (human ruling, 2026-07-12
# — "no need for large tiff files"). admit() converts single-frame tiffs to
# PNG in place and the PNG becomes the original; TIFF_EXTS therefore stays
# out of IMAGE_EXTS — it is admit-time-special, not a plain pass-through
# image format, and (unlike IMAGE_EXTS) never appears in sources/.
TIFF_EXTS = {".tif", ".tiff"}
MAIL_EXTS = {".eml"}
TEXT_EXTS = {".md", ".txt", ".json", ".yaml", ".yml"}
SUPPORTED_EXTS = (EXTRACTIVE_EXTS | LIBREOFFICE_EXTS | IMAGE_EXTS | TIFF_EXTS
                  | MAIL_EXTS | TEXT_EXTS | {".pdf"})

# Containers are DELIVERY VEHICLES, not originals: `unpack` extracts them
# into inbox/ and discards the zip — the members become the originals.
# They must never join SUPPORTED_EXTS; admit keeps refusing them.
CONTAINER_EXTS = {".zip", ".athena"}

# Zones a `link` doc path may point into — inbox/sources/workbench never
# hold finished documents, only originals and ephemeral staging.
DOC_ZONES = {"library", "activities", "output"}

TEXT_COVERAGE_MIN_CHARS = 100   # below this a page is scanned/figure
RENDER_DPI = 150
EXCEL_ERRORS = {"#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?", "#NULL!",
                "#NUM!"}

TODAY = datetime.now().strftime("%Y-%m-%d")


class IntakeError(Exception):
    """Strict-writer refusal — named loudly, never silently skipped."""


# ── shared helpers ───────────────────────────────────────────────────────

def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


_PNG_NATIVE_MODES = {"1", "L", "LA", "P", "RGB", "RGBA", "I"}


def tiff_to_png(src: Path, dest: Path) -> None:
    """Deterministic, lossless-pixel tiff -> PNG. Single frame only: a
    multi-frame tiff (fax/scan stacks, layered exports) is refused outright
    rather than silently keeping just the first frame — the fidelity
    contract (SKILL.md: "a lossy conversion is a failed conversion") rules
    out a silent partial admit; the conservative choice is refusal, not a
    loud-warning first-frame-wins."""
    from PIL import Image

    with Image.open(src) as img:
        n_frames = getattr(img, "n_frames", 1)
        if n_frames > 1:
            raise IntakeError(
                f"multi-frame tiff refused: {src.name} carries {n_frames} "
                "frames — fusion-intake admits single-frame tiffs only "
                "(silently keeping just the first frame would violate the "
                "fidelity contract); split it into single-frame files and "
                "admit them individually, or convert it to a single-frame "
                "tiff/png before dropping it in inbox/")
        img.load()
        # PNG cannot encode every tiff pixel mode (CMYK, palette variants
        # with exotic depth, …) — fall back to a standard mode rather than
        # let Pillow's save() raise. RGB/RGBA/L/LA/P/1/I pass through with
        # their pixels untouched (lossless); anything else is converted
        # once, to the closest PNG-native mode (RGBA if it carries alpha,
        # else RGB).
        if img.mode not in _PNG_NATIVE_MODES:
            img = img.convert("RGBA" if "A" in img.mode else "RGB")
        dest.parent.mkdir(parents=True, exist_ok=True)
        img.save(dest, format="PNG")


def slugify(name: str) -> str:
    """SPEC §4 filename: lowercase, hyphen-separated, stem <=60 chars."""
    stem = re.sub(r"\.(xlsx|docx|pptx|csv|pdf|eml|md|txt|png|jpe?g|webp|gif|html?|json|ya?ml)$",
                  "", name, flags=re.IGNORECASE)
    stem = stem.lower()
    stem = re.sub(r"[^a-z0-9]+", "-", stem).strip("-")
    stem = re.sub(r"-+", "-", stem)
    return stem[:60].rstrip("-") or "document"


def cell_to_string(cell) -> str:
    if isinstance(cell, str) and cell.strip() in EXCEL_ERRORS:
        return ""
    if cell is None:
        return ""
    if isinstance(cell, bool):
        return str(cell)
    if isinstance(cell, float):
        if cell.is_integer():
            return str(int(cell))
        return repr(cell)   # shortest round-trippable form — verbatim, never rounded
    if isinstance(cell, int):
        return str(cell)
    if isinstance(cell, datetime):
        return cell.strftime("%Y-%m-%d")
    text = str(cell).replace("|", "\\|").replace("\n", "<br>")
    return text.strip()


def _row_empty(row) -> bool:
    return all(cell_to_string(c) == "" for c in row)


def prune_empty_columns(rows):
    if not rows:
        return rows
    width = max(len(r) for r in rows)
    padded = [list(r) + [None] * (width - len(r)) for r in rows]
    keep = [i for i in range(width)
            if any(cell_to_string(r[i]) != "" for r in padded)]
    return [[r[i] for i in keep] for r in padded] if keep else []


def rows_to_table(rows) -> str:
    """One markdown table. Every row, every column — no caps, no sampling."""
    rows = [r for r in rows if not _row_empty(r)]
    rows = prune_empty_columns(rows)
    if not rows or not rows[0]:
        return "*No data*"
    string_rows = [[cell_to_string(c) for c in r] for r in rows]
    lines = ["| " + " | ".join(string_rows[0]) + " |",
             "| " + " | ".join(["---"] * len(rows[0])) + " |"]
    lines += ["| " + " | ".join(r) + " |" for r in string_rows[1:]]
    return "\n".join(lines)


def render_document(fm: dict, summary: str, body: str) -> str:
    front = yaml.safe_dump(fm, sort_keys=False, allow_unicode=True,
                           width=2**31 - 1)
    return f"---\n{front}---\n\n## Summary\n\n{summary}\n\n---\n\n{body}\n"


# ── MANIFEST (single writer lives here) ──────────────────────────────────

def _manifest_path(root: Path) -> Path:
    return root / "sources" / "MANIFEST.md"


def manifest_append(root: Path, rel: str, actor: str, sha: str) -> None:
    path = _manifest_path(root)
    if not path.is_file():
        path.write_text("# Manifest\n\n| file | added | by | sha256 | library |\n"
                        "|---|---|---|---|---|\n", encoding="utf-8", newline="\n")
    text = path.read_text(encoding="utf-8")
    if not text.endswith("\n"):
        text += "\n"
    text += f"| {rel} | {TODAY} | {actor} | {sha} | — |\n"
    path.write_text(text, encoding="utf-8", newline="\n")


def manifest_link(root: Path, rel: str, doc: str) -> None:
    path = _manifest_path(root)
    if not path.is_file():
        raise IntakeError(f"no manifest at {path}")
    lines = path.read_text(encoding="utf-8").splitlines()
    hit = False
    for i, line in enumerate(lines):
        if not line.strip().startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) == 5 and cells[0] == rel:
            cells[4] = doc if cells[4] in ("—", "") else f"{cells[4]}, {doc}"
            lines[i] = "| " + " | ".join(cells) + " |"
            hit = True
            break
    if not hit:
        raise IntakeError(f"source not in manifest: {rel}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")


def manifest_relink(root: Path, rel: str, old_doc: str, new_doc: str) -> None:
    path = _manifest_path(root)
    if not path.is_file():
        raise IntakeError(f"no manifest at {path}")
    if old_doc == new_doc:
        raise IntakeError(
            f"relink is a change: --from and --to are both {old_doc!r}")
    if not (Path(root) / new_doc).is_file():
        raise IntakeError(f"relink target does not exist on disk: {new_doc}")
    lines = path.read_text(encoding="utf-8").splitlines()
    hit = False
    for i, line in enumerate(lines):
        if not line.strip().startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) == 5 and cells[0] == rel:
            values = ([] if cells[4] in ("—", "")
                      else [v.strip() for v in cells[4].split(",")])
            if old_doc not in values:
                raise IntakeError(
                    f"cell does not carry {old_doc!r} for source {rel} "
                    f"(cell: {cells[4]!r})")
            if new_doc in values:
                raise IntakeError(
                    f"cell already carries {new_doc!r} for source {rel}")
            cells[4] = ", ".join(
                new_doc if v == old_doc else v for v in values)
            lines[i] = "| " + " | ".join(cells) + " |"
            hit = True
            break
    if not hit:
        raise IntakeError(f"source not in manifest: {rel}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8", newline="\n")


def _manifest_rels(root: Path) -> set:
    """The `file` column of every real row — used by `batch` to validate
    (category, basename) uniqueness against what's already registered."""
    path = _manifest_path(root)
    if not path.is_file():
        return set()
    rels = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip().startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) == 5 and cells[0] not in ("file", "---"):
            rels.add(cells[0])
    return rels


# ── admit ────────────────────────────────────────────────────────────────

def admit(root: Path, inbox_rel: str, category: str, actor: str) -> dict:
    root = Path(root)
    # Resolve the inbox root once and require the target to stay under it —
    # inbox_rel is untrusted (may carry ../ traversal) and a LEXICAL
    # relative_to() check downstream would accept a resolved parent that has
    # already escaped. Check this BEFORE any other validation, including
    # is_file() (mirrors unpack()'s pattern, commit 8851fd3) — admit is the
    # only writer of sources/MANIFEST.md and every caller (batch included)
    # goes through this one function, so the guard lives here once.
    inbox_root = (root / "inbox").resolve()
    src = root / "inbox" / inbox_rel
    src_resolved = src.resolve()
    if not src_resolved.is_relative_to(inbox_root):
        raise IntakeError(f"inbox path escapes inbox/: {inbox_rel!r}")
    if not src.is_file():
        raise IntakeError(f"not in inbox/: {inbox_rel}")
    if src.suffix.lower() not in SUPPORTED_EXTS:
        raise IntakeError(
            f"unsupported format: {src.suffix.lower()} — the gate refuses "
            "what it cannot preserve; the file stays in inbox/")
    if not actor or any(c.isspace() for c in actor):
        raise IntakeError(f"actor must be a single token: {actor!r}")
    category = category.strip().strip("/")
    if not category:
        raise IntakeError("category required")
    if any(c in src.name for c in "|\n\r") or any(c in category for c in "|\n\r"):
        raise IntakeError(
            f"'|' and newlines break the manifest grammar: {src.name!r} — "
            "rename the incoming file before admitting it")
    is_tiff = src.suffix.lower() in TIFF_EXTS
    # A tiff is never admitted as a tiff (2026-07-12 ruling): the PNG IS the
    # original from here on, so the collision check, the MANIFEST row, and
    # the returned `source` all name the .png — never the discarded .tif.
    dest_name = f"{src.stem}.png" if is_tiff else src.name
    dest = root / "sources" / category / dest_name
    if dest.exists():
        raise IntakeError(
            f"sources/ is immutable and {category}/{dest_name} already "
            "exists — rename the incoming file before admitting it")
    if is_tiff:
        original_name = src.name
        original_sha = sha256_of(src)
        tiff_to_png(src, dest)          # raises before touching dest on
        src.unlink()                    # multi-frame; conservative refusal
        sha = sha256_of(dest)
    else:
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(src), str(dest))
        sha = sha256_of(dest)
    rel = f"{category}/{dest_name}"
    manifest_append(root, rel, actor, sha)
    result = {"source": rel, "sha256": sha, "manifest_row": True}
    if is_tiff:
        # Ledger-visible honesty: the MANIFEST row only ever names the PNG
        # (the admitted original), so the tiff's identity is carried here
        # for the caller to sign — SKILL.md step 2 logs this `note` via
        # `fusion log noted` right after admit, same run.
        result["converted"] = {
            "from_format": "tiff", "to_format": "png",
            "original_name": original_name, "original_sha256": original_sha,
        }
        result["note"] = (
            f"tiff→png at admit: {original_name} (sha256 {original_sha}) "
            f"→ {dest_name} (sha256 {sha}) — original tiff bytes not "
            "retained (waived per 2026-07-12 ruling: no need for large "
            "tiff files)")
    return result


# ── unpack: containers are delivery vehicles, not originals ─────────────

def unpack(root: Path, inbox_rel: str) -> dict:
    """Extract a container (.zip / .athena) sitting in inbox/ into a sibling
    folder — the members become the originals, the container is discarded.
    Does NOT write the ledger: the operator signs a `noted` entry via
    `fusion log` (single-writer)."""
    root = Path(root)
    # Resolve the inbox root once and require every path we touch to stay
    # under it — inbox_rel is untrusted (may carry ../ traversal) and a
    # LEXICAL relative_to() check downstream would accept a resolved parent
    # that has already escaped. Check this BEFORE any other validation.
    inbox_root = (root / "inbox").resolve()
    src = root / "inbox" / inbox_rel
    src_resolved = src.resolve()
    if not src_resolved.is_relative_to(inbox_root):
        raise IntakeError(
            f"container path escapes inbox/: {inbox_rel!r}")
    if not src.is_file():
        raise IntakeError(f"not in inbox/: {inbox_rel}")
    if src.suffix.lower() not in CONTAINER_EXTS:
        raise IntakeError(
            f"not a container: {src.suffix.lower()} — unpack only handles "
            f"{sorted(CONTAINER_EXTS)}")
    if not zipfile.is_zipfile(src):
        raise IntakeError(f"not a readable zip: {inbox_rel}")

    stem = src_resolved.stem
    dest = src_resolved.parent / stem   # sibling of the container — nested
                                # vehicles keep their folder context, not
                                # inbox root — computed from the RESOLVED
                                # src so a traversal-crafted parent can't
                                # smuggle a dest outside inbox/ either.
    if not dest.is_relative_to(inbox_root):
        raise IntakeError(
            f"container path escapes inbox/: {inbox_rel!r}")
    dest_rel = str(dest.relative_to(inbox_root.parent))
    if dest.exists():
        raise IntakeError(
            f"destination already exists: {dest_rel} — refusing to "
            "merge a container into existing content")

    with zipfile.ZipFile(src) as zf:
        members = []
        for info in zf.infolist():
            if info.is_dir():
                continue
            name = info.filename
            if name.startswith("__MACOSX/") or Path(name).name.startswith("._"):
                continue
            members.append(info)

        # Validate every destination BEFORE writing anything — a hostile
        # member fails the whole unpack, not just itself.
        dest_resolved = dest.resolve()
        targets = []
        for info in members:
            target = (dest / info.filename).resolve()
            if not target.is_relative_to(dest_resolved):
                raise IntakeError(
                    f"zip-slip attempt in {inbox_rel}: hostile member "
                    f"{info.filename!r} escapes {dest_rel}/ — unpack refused")
            targets.append((info, target))

        dest.mkdir(parents=True)
        try:
            for info, target in targets:
                target.parent.mkdir(parents=True, exist_ok=True)
                with zf.open(info) as fh, open(target, "wb") as out:
                    shutil.copyfileobj(fh, out)
        except Exception:
            shutil.rmtree(dest, ignore_errors=True)
            raise

    src.unlink()
    return {"unpacked": len(targets), "dir": dest_rel}


# ── prepare: routing + engines ───────────────────────────────────────────

def probe_pdf_text_layer(pdf_path: Path):
    import fitz
    records = []
    doc = fitz.open(str(pdf_path))
    try:
        for i, page in enumerate(doc, 1):
            text = page.get_text().strip()
            records.append({"page": i, "text": text, "text_chars": len(text),
                            "needs_vision": len(text) < TEXT_COVERAGE_MIN_CHARS})
    finally:
        doc.close()
    return records


def render_pdf_pages(pdf_path: Path, out_dir: Path, pages=None):
    import fitz
    out_dir.mkdir(parents=True, exist_ok=True)
    written = []
    doc = fitz.open(str(pdf_path))
    try:
        targets = (list(range(1, doc.page_count + 1))
                   if pages is None else sorted(pages))
        for n in targets:
            if 1 <= n <= doc.page_count:
                pix = doc.load_page(n - 1).get_pixmap(dpi=RENDER_DPI)
                img = out_dir / f"page-{n:03d}.png"
                pix.save(str(img))
                written.append(img)
    finally:
        doc.close()
    return written


def require_soffice() -> str:
    exe = shutil.which("soffice") or shutil.which("libreoffice")
    if not exe:
        raise IntakeError(
            "LibreOffice not found: 'soffice' must be on PATH for "
            "docx/pptx/legacy office formats (declared in SKILL.md "
            "compatibility). Install it or add it to PATH. No fallback.")
    return exe


def soffice_to_pdf(src: Path, out_dir: Path) -> Path:
    exe = require_soffice()
    out_dir.mkdir(parents=True, exist_ok=True)
    proc = subprocess.run(
        [exe, "--headless", "--convert-to", "pdf", "--outdir",
         str(out_dir), str(src)],
        capture_output=True, text=True, timeout=300)
    pdf = out_dir / (src.stem + ".pdf")
    if proc.returncode != 0 or not pdf.exists():
        raise IntakeError(
            f"soffice failed on {src.name} (rc={proc.returncode}): "
            f"{proc.stderr.strip()[:400]}")
    return pdf


class _HTMLText(HTMLParser):
    """Text of an HTML mail body: entities decoded (convert_charrefs),
    script/style dropped whole, block tags become line breaks."""
    _SKIP = {"script", "style"}
    _BREAK = {"p", "br", "div", "tr", "li", "table",
              "h1", "h2", "h3", "h4", "h5", "h6"}

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self._chunks: list = []
        self._skip_depth = 0

    def handle_starttag(self, tag, attrs):
        if tag in self._SKIP:
            self._skip_depth += 1
        elif tag in self._BREAK:
            self._chunks.append("\n")

    def handle_endtag(self, tag):
        if tag in self._SKIP and self._skip_depth:
            self._skip_depth -= 1
        elif tag in self._BREAK:
            self._chunks.append("\n")

    def handle_data(self, data):
        if not self._skip_depth:
            self._chunks.append(data)


def html_to_text(html: str) -> str:
    parser = _HTMLText()
    parser.feed(html)
    parser.close()
    lines = [re.sub(r"[ \t]+", " ", ln).strip()
             for ln in "".join(parser._chunks).splitlines()]
    return "\n".join(ln for ln in lines if ln)


def eml_to_text(path: Path, work_dir: Path):
    msg = BytesParser(policy=policy.default).parse(open(path, "rb"))
    lines = [f"{h}: {msg[h]}" for h in ("From", "To", "Date", "Subject")
             if msg[h]]
    body = msg.get_body(preferencelist=("plain", "html"))
    content = body.get_content() if body else ""
    if body and body.get_content_subtype() == "html":
        content = html_to_text(content)
    attachments = []
    seen_lower = set()
    for part in msg.iter_attachments():
        name = Path(part.get_filename() or "attachment.bin").name
        if name in {"", ".", ".."}:
            name = "attachment.bin"
        stem, dot, ext = name.partition(".")
        n = 2
        while name.lower() in seen_lower:
            name = f"{stem}-{n}{dot}{ext}"
            n += 1
        (work_dir / name).write_bytes(part.get_payload(decode=True) or b"")
        attachments.append(name)
        seen_lower.add(name.lower())
    return "\n".join(lines) + "\n\n" + content, attachments


def _route(ext: str) -> str:
    ext = ext.lower()
    if ext in EXTRACTIVE_EXTS:
        return "extractive"
    if ext in LIBREOFFICE_EXTS:
        return "libreoffice"
    if ext == ".pdf":
        return "pdf"
    if ext in IMAGE_EXTS:
        return "image"
    if ext in MAIL_EXTS:
        return "mail"
    if ext in TEXT_EXTS:
        return "text"
    raise IntakeError(f"unsupported format: {ext} — the gate refuses "
                      "what it cannot preserve")


def _sheet_matrix(ws):
    """All cell values with merged ranges unfolded — every cell a merge
    spans carries the anchor's value, so pruning never eats merged data."""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    for rng in ws.merged_cells.ranges:
        if rng.min_row - 1 >= len(rows):
            continue
        anchor_row = rows[rng.min_row - 1]
        anchor = (anchor_row[rng.min_col - 1]
                  if rng.min_col - 1 < len(anchor_row) else None)
        for r in range(rng.min_row, min(rng.max_row, len(rows)) + 1):
            row = rows[r - 1]
            for c in range(rng.min_col, min(rng.max_col, len(row)) + 1):
                row[c - 1] = anchor
    return rows


def _xlsx_body(path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    sections, sheets, rows_total = [], 0, 0
    for name in wb.sheetnames:
        rows = _sheet_matrix(wb[name])
        live = [r for r in rows if not _row_empty(r)]
        if live:
            sheets += 1
            rows_total += max(len(live) - 1, 0)
            sections.append(f"## {name}\n\n{rows_to_table(rows)}")
    return "\n\n".join(sections) or "*No data*", sheets, rows_total


def _csv_body(path: Path):
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        rows = list(csv_mod.reader(fh))
    return rows_to_table(rows), 1, max(len(rows) - 1, 0)


def _work_dir(root: Path) -> Path:
    d = root / "workbench" / ".intake" / uuid.uuid4().hex[:12]
    d.mkdir(parents=True, exist_ok=True)
    return d


def _read_frontmatter(path: Path) -> dict:
    """Tolerant frontmatter reader: {} if the leading --- block is absent
    or unparsable — reconcile must never choke on a hand-edited document."""
    text = path.read_text(encoding="utf-8")
    if not text.startswith("---\n"):
        return {}
    parts = text.split("---\n", 2)
    if len(parts) < 3:
        return {}
    try:
        fm = yaml.safe_load(parts[1])
    except yaml.YAMLError:
        return {}
    return fm if isinstance(fm, dict) else {}


def _merge_reconcile_fm(existing_fm: dict, seed: dict) -> dict:
    """The librarian curated the existing document — its values win.
    Existing keys come first (unknown keys preserved verbatim, including
    the original `created:`); title/type/aurora are filled from the seed
    only where the existing doc lacks them; `source:` always repoints to
    the new original; `updated:` is always bumped to today."""
    merged = dict(existing_fm)
    for key in ("title", "type", "aurora"):
        if not merged.get(key):
            merged[key] = seed[key]
    if "created" not in merged:
        merged["created"] = seed["created"]
    merged["source"] = seed["source"]
    merged["updated"] = TODAY
    return merged


def prepare(root: Path, source_rel: str, dest: str | None = None,
            slug: str | None = None, doc_type: str | None = None,
            aurora: str = "library", reconcile: bool = False) -> dict:
    root = Path(root)
    src = root / "sources" / source_rel
    if not src.is_file():
        raise IntakeError(f"not in sources/: {source_rel}")
    if aurora not in AURORAS:
        raise IntakeError(f"aurora must be one of the eight, got: {aurora!r}")
    category = Path(source_rel).parts[0] if "/" in source_rel else "reference"
    doc_type = doc_type or category.rstrip("s")
    dest = (dest or f"library/{category}").strip("/")
    slug = slug or slugify(src.name)
    out_rel = f"{dest}/{slug}.md"
    seed = {"title": src.stem, "type": doc_type, "aurora": aurora,
            "source": f"sources/{source_rel}", "created": TODAY}
    path = _route(src.suffix)

    if path == "extractive":
        if src.suffix.lower() == ".xlsx":
            body, sheets, nrows = _xlsx_body(src)
        else:
            body, sheets, nrows = _csv_body(src)
        summary = (f"Tabular data converted from {src.name}: "
                   f"{sheets} sheet(s), {nrows} data row(s).")
        out = root / out_rel
        if out.exists() and not reconcile:
            raise IntakeError(
                f"document exists: {out_rel} — pass --reconcile for a "
                "confirmed update, or choose --slug")
        if reconcile and out.exists():
            seed = _merge_reconcile_fm(_read_frontmatter(out), seed)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(render_document(seed, summary, body), encoding="utf-8",
                       newline="\n")
        return {"path": "extractive", "done": True, "source": source_rel,
                "output_file": out_rel, "front_matter_seed": seed,
                "reconcile": reconcile}

    # vision / structured paths get a work dir + manifest for Stage 2
    out = root / out_rel
    if out.exists() and not reconcile:
        raise IntakeError(
            f"document exists: {out_rel} — pass --reconcile for a "
            "confirmed update, or choose --slug")
    if reconcile and out.exists():
        seed = _merge_reconcile_fm(_read_frontmatter(out), seed)

    work = _work_dir(root)
    record = {"path": path, "done": False, "source": source_rel,
              "run_dir": str(work.relative_to(root)),
              "output_file": out_rel, "front_matter_seed": seed,
              "reconcile": reconcile,
              "pages": [], "images": [], "attachments": [],
              "intermediate_pdf": None}

    if path == "libreoffice":
        pdf = soffice_to_pdf(src, work)
        record["intermediate_pdf"] = str(pdf.relative_to(root))
        record["pages"] = probe_pdf_text_layer(pdf)
        imgs = render_pdf_pages(pdf, work)                    # ALL pages
        record["images"] = [str(p.relative_to(root)) for p in imgs]
    elif path == "pdf":
        probe = probe_pdf_text_layer(src)
        record["pages"] = probe
        if all(p["needs_vision"] for p in probe):
            record["path"] = "pdf_scanned"
            imgs = render_pdf_pages(src, work)                # ALL pages
        else:
            record["path"] = "pdf_text"
            imgs = render_pdf_pages(
                src, work, pages=[p["page"] for p in probe if p["needs_vision"]])
        record["images"] = [str(p.relative_to(root)) for p in imgs]
    elif path == "image":
        copy = work / src.name
        shutil.copy2(src, copy)
        record["pages"] = [{"page": 1, "text": "", "text_chars": 0,
                            "needs_vision": True}]
        record["images"] = [str(copy.relative_to(root))]
    elif path == "mail":
        text, attachments = eml_to_text(src, work)
        record["pages"] = [{"page": 1, "text": text,
                            "text_chars": len(text), "needs_vision": False}]
        record["attachments"] = attachments
    else:  # text passthrough
        text = src.read_text(encoding="utf-8", errors="replace")
        record["pages"] = [{"page": 1, "text": text,
                            "text_chars": len(text), "needs_vision": False}]

    record["page_count"] = len(record["pages"])
    manifest = work / "manifest.json"
    manifest.write_text(json.dumps(record, indent=2), encoding="utf-8",
                        newline="\n")
    record["manifest"] = str(manifest.relative_to(root))
    return record


def link(root: Path, source_rel: str, doc_rel: str) -> dict:
    manifest_link(Path(root), source_rel, doc_rel)
    return {"source": source_rel, "library": doc_rel}


def relink(root: Path, source_rel: str, old_doc: str, new_doc: str) -> dict:
    manifest_relink(Path(root), source_rel, old_doc, new_doc)
    return {"source": source_rel, "from": old_doc, "to": new_doc}


# ── batch: one op-list, validated whole, then moved ──────────────────────
#
# Schema (references/delivery.md):
#   {"admits": [{"file": "<inbox-rel>", "category": "<cat>"}],
#    "links":  [{"source": "<sources-rel>", "doc": "<zone-rel-doc>"}]}
#
# Semantics — two validation passes, never partial:
#   1. VALIDATE every admit and link syntactically/structurally BEFORE any
#      filesystem or MANIFEST mutation (the gate's rule, batch-scale): a
#      bad op anywhere in the list means nothing moves, nothing is
#      appended — not even the good ops ahead of it.
#   2. EXECUTE all admits (reusing `admit()`, one MANIFEST row per file).
#   3. VALIDATE that every link's doc exists on disk NOW — admits already
#      landed (the mechanical half doesn't roll back), but if any doc is
#      missing NO link is written, not even the ones whose doc is fine.
#   4. EXECUTE all links (reusing `link()`).

def batch(root: Path, ops: dict, actor: str) -> dict:
    root = Path(root)
    if not actor or any(c.isspace() for c in actor):
        raise IntakeError(f"actor must be a single token: {actor!r}")

    admits = ops.get("admits", [])
    links = ops.get("links", [])

    # ── phase 1: validate every op, nothing touched yet ──────────────────
    existing_rels = _manifest_rels(root)
    batch_rels = set()
    seen_files = set()
    for i, op in enumerate(admits):
        file = (op.get("file") or "").strip()
        category = (op.get("category") or "").strip().strip("/")
        if not file:
            raise IntakeError(f"admits[{i}]: file required")
        if not category:
            raise IntakeError(f"admits[{i}]: category required")
        if file in seen_files:
            raise IntakeError(
                f"admits[{i}]: {file!r} is admitted twice in this batch — "
                "the first admit would move it before the second runs")
        seen_files.add(file)
        src = root / "inbox" / file
        if not src.is_file():
            raise IntakeError(f"admits[{i}]: not in inbox/: {file}")
        if src.suffix.lower() not in SUPPORTED_EXTS:
            raise IntakeError(
                f"admits[{i}]: unsupported format: {src.suffix.lower()} — "
                "the gate refuses what it cannot preserve; the file stays "
                "in inbox/")
        if any(c in src.name for c in "|\n\r") or any(c in category for c in "|\n\r"):
            raise IntakeError(
                f"admits[{i}]: '|' and newlines break the manifest "
                f"grammar: {src.name!r}")
        rel = f"{category}/{src.name}"
        # A hand-placed file in sources/ with no MANIFEST row is a collision
        # too — checking existing_rels alone lets a mixed batch admit past
        # it, then fail mid-batch at admit()'s own dest.exists() with no
        # rollback, breaking the all-or-nothing promise (references/delivery.md).
        if rel in existing_rels or (root / "sources" / rel).exists():
            raise IntakeError(
                f"admits[{i}]: sources/ is immutable and {rel} already "
                "exists — rename the incoming file before admitting it")
        if rel in batch_rels:
            raise IntakeError(
                f"admits[{i}]: duplicate (category, basename) within this "
                f"batch: {rel}")
        batch_rels.add(rel)

    for i, op in enumerate(links):
        source = (op.get("source") or "").strip()
        doc = (op.get("doc") or "").strip()
        if not source:
            raise IntakeError(f"links[{i}]: source required")
        if not doc:
            raise IntakeError(f"links[{i}]: doc required")
        # Normalize before judging the zone — a LEXICAL split on the raw
        # string ("library/../inbox/evil.md".split("/",1)[0] == "library")
        # passes a path that actually resolves outside every zone. Collapse
        # ".." segments first, then require: no leftover ".." (it climbed
        # above the zone root), not absolute, first segment in DOC_ZONES.
        norm_doc = posixpath.normpath(doc)
        doc_parts = norm_doc.split("/")
        if (posixpath.isabs(norm_doc) or doc_parts[0] == ".."
                or doc_parts[0] not in DOC_ZONES):
            raise IntakeError(
                f"links[{i}]: doc must be zone-relative under "
                f"{sorted(DOC_ZONES)}: {doc!r}")
        op["doc"] = norm_doc   # store/compare the NORMALIZED path — later
                                # phases (doc-exists check, link()) read
                                # this same op dict, never the raw input
        if source not in existing_rels and source not in batch_rels:
            raise IntakeError(
                f"links[{i}]: link source will not exist: {source!r} — "
                "not already registered in sources/ nor admitted by this "
                "batch's own admits")

    # ── phase 2: execute admits ───────────────────────────────────────────
    admitted = [admit(root, op["file"], op["category"], actor) for op in admits]

    # ── phase 3: validate every link's doc exists NOW, before any link ───
    for i, op in enumerate(links):
        doc = op["doc"].strip()
        if not (root / doc).is_file():
            raise IntakeError(
                f"links[{i}]: doc does not exist: {doc!r} — write it "
                "(Stage 2 conversion) before linking; no link in this "
                "batch was written")

    # ── phase 4: execute links ────────────────────────────────────────────
    for op in links:
        link(root, op["source"].strip(), op["doc"].strip())

    return {"admitted": len(admitted), "linked": len(links)}


def cleanup(run_dir: Path) -> None:
    run_dir = Path(run_dir).resolve()
    parts = run_dir.parts
    if ".intake" not in parts or "workbench" not in parts:
        raise IntakeError(f"refusing to delete outside workbench/.intake: {run_dir}")
    if run_dir.exists():
        shutil.rmtree(run_dir)


# ── CLI ──────────────────────────────────────────────────────────────────

def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="fusion-intake Stage-1 engine")
    sub = ap.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("admit", help="inbox -> sources + MANIFEST row")
    p.add_argument("--bucket", required=True)
    p.add_argument("--file", required=True, help="path relative to inbox/")
    p.add_argument("--category", required=True)
    p.add_argument("--actor", required=True)

    p = sub.add_parser("unpack", help="container -> inbox folder "
                       "(vehicle, not an original)")
    p.add_argument("--bucket", required=True)
    p.add_argument("--file", required=True, help="path relative to inbox/")

    p = sub.add_parser("prepare", help="route + convert / stage for vision")
    p.add_argument("--bucket", required=True)
    p.add_argument("--source", required=True, help="path relative to sources/")
    p.add_argument("--dest", help="zone-relative output dir "
                   "(default library/<category>)")
    p.add_argument("--slug")
    p.add_argument("--type", dest="doc_type")
    p.add_argument("--aurora", default="library")
    p.add_argument("--reconcile", action="store_true",
                   help="confirmed update: reconcile the existing document "
                   "in place instead of refusing the slug collision")

    p = sub.add_parser("link", help="set the MANIFEST library column")
    p.add_argument("--bucket", required=True)
    p.add_argument("--source", required=True)
    p.add_argument("--doc", required=True)

    p = sub.add_parser("relink", help="repoint one value in a MANIFEST "
                       "library cell after a document moved")
    p.add_argument("--bucket", required=True)
    p.add_argument("--source", required=True,
                   help="path relative to sources/")
    p.add_argument("--from", dest="from_doc", required=True,
                   help="current zone-relative doc path in the cell")
    p.add_argument("--to", dest="to_doc", required=True,
                   help="new zone-relative doc path (must exist)")

    p = sub.add_parser("cleanup", help="delete one work dir")
    p.add_argument("--run-dir", required=True)

    p = sub.add_parser("batch", help="validated multi-op admit+link in one "
                       "process (references/delivery.md)")
    p.add_argument("--bucket", required=True)
    p.add_argument("--ops", required=True,
                   help="path to a JSON op-list: {admits: [...], links: [...]}")
    p.add_argument("--actor", default="claude")

    args = ap.parse_args(argv)
    try:
        if args.cmd == "admit":
            out = admit(Path(args.bucket), args.file, args.category, args.actor)
        elif args.cmd == "unpack":
            out = unpack(Path(args.bucket), args.file)
        elif args.cmd == "prepare":
            out = prepare(Path(args.bucket), args.source, dest=args.dest,
                          slug=args.slug, doc_type=args.doc_type,
                          aurora=args.aurora, reconcile=args.reconcile)
        elif args.cmd == "link":
            out = link(Path(args.bucket), args.source, args.doc)
        elif args.cmd == "relink":
            out = relink(Path(args.bucket), args.source,
                         args.from_doc, args.to_doc)
        elif args.cmd == "batch":
            ops = json.loads(Path(args.ops).read_text(encoding="utf-8"))
            out = batch(Path(args.bucket), ops, args.actor)
        else:
            cleanup(Path(args.run_dir))
            out = {"cleaned": args.run_dir}
    except IntakeError as exc:
        print(f"intake: {exc}", file=sys.stderr)
        return 1
    print(json.dumps(out, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
