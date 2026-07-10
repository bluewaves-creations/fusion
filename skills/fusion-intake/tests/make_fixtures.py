#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1.0", "pymupdf>=1.24.0"]
# ///
"""Generate real legacy-format fixtures for the intake tests. Imported by
the test modules; also runnable standalone to eyeball the artifacts."""
import zipfile
from email.message import EmailMessage
from pathlib import Path


def make_xlsx(path: Path):
    """Two sheets, exact numbers, a pipe cell, an all-empty column."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scores"
    ws.append(["supplier", "score", None, "notes"])   # 3rd column all-empty
    ws.append(["Acme Corp", 78.5, None, "steady | improving"])
    ws.append(["Northwind", 91, None, "top tier"])
    ws2 = wb.create_sheet("Meta")
    ws2.append(["generated", "2026-07-01"])
    wb.save(path)


def make_csv(path: Path):
    path.write_bytes(
        ("﻿" "item,qty,price\nJazzmaster 1962,1,12500.50\n"
         "Pedalboard,1,2300\n").encode("utf-8"))


def make_text_pdf(path: Path):
    """Two pages with a healthy text layer."""
    import fitz
    doc = fitz.open()
    for n in (1, 2):
        page = doc.new_page()
        text = (f"Page {n} of the supplier audit. " * 12)
        page.insert_text((72, 72), text, fontsize=11)
    doc.save(path)
    doc.close()


def make_scanned_pdf(path: Path):
    """One page, no text layer at all — drawing only."""
    import fitz
    doc = fitz.open()
    page = doc.new_page()
    page.draw_line((72, 72), (400, 400))
    doc.save(path)
    doc.close()


DOCX_DOCUMENT = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
<w:body>
<w:p><w:r><w:t>Onboarding procedure for the studio bucket.</w:t></w:r></w:p>
<w:p><w:r><w:t>Step one: read the conventions. Step two: sign the ledger.</w:t></w:r></w:p>
</w:body></w:document>"""

DOCX_CONTENT_TYPES = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>"""

DOCX_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>"""


def make_docx(path: Path):
    """Minimal valid OOXML — LibreOffice opens it happily."""
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", DOCX_CONTENT_TYPES)
        z.writestr("_rels/.rels", DOCX_RELS)
        z.writestr("word/document.xml", DOCX_DOCUMENT)


def make_eml(path: Path):
    msg = EmailMessage()
    msg["From"] = "gilberte@example.com"
    msg["To"] = "studio@example.com"
    msg["Subject"] = "Re: rehearsal schedule"
    msg["Date"] = "Thu, 09 Jul 2026 10:15:00 +0200"
    msg.set_content("Rehearsal moves to Thursday. Bring the Jazzmaster.\n")
    msg.add_attachment(b"setlist,song\n1,First Light\n",
                       maintype="text", subtype="csv", filename="setlist.csv")
    path.write_bytes(bytes(msg))


def make_png(path: Path):
    """Tiny valid PNG (1x1, red) — hand-assembled, no dependencies."""
    import struct, zlib
    def chunk(tag, data):
        c = tag + data
        return struct.pack(">I", len(data)) + c + struct.pack(
            ">I", zlib.crc32(c) & 0xFFFFFFFF)
    ihdr = struct.pack(">IIBBBBB", 1, 1, 8, 2, 0, 0, 0)
    idat = zlib.compress(b"\x00\xff\x00\x00")
    path.write_bytes(b"\x89PNG\r\n\x1a\n" + chunk(b"IHDR", ihdr)
                     + chunk(b"IDAT", idat) + chunk(b"IEND", b""))


DOCX_TWO_PAGE_DOCUMENT = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
<w:body>
<w:p><w:r><w:t>Page one of the onboarding pack.</w:t></w:r></w:p>
<w:p><w:r><w:br w:type="page"/></w:r></w:p>
<w:p><w:r><w:t>Page two: the ledger discipline in daily practice.</w:t></w:r></w:p>
</w:body></w:document>"""


def make_docx_two_page(path: Path):
    """Two pages forced by an explicit page break."""
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", DOCX_CONTENT_TYPES)
        z.writestr("_rels/.rels", DOCX_RELS)
        z.writestr("word/document.xml", DOCX_TWO_PAGE_DOCUMENT)


def make_html_eml(path: Path):
    """HTML-only mail: entities, style and script blocks, no text/plain part."""
    msg = EmailMessage()
    msg["From"] = "gilberte@example.com"
    msg["To"] = "studio@example.com"
    msg["Subject"] = "Studio move — final quote"
    msg["Date"] = "Fri, 10 Jul 2026 09:00:00 +0200"
    msg.set_content(
        "<html><head><style>p{color:red}</style>"
        "<script>alert('tracking')</script></head>"
        "<body><p>Quote for Dupont &amp; Fils: total 1250 euros.</p>"
        "<p>Valid until <b>2026-08-01</b>.</p></body></html>",
        subtype="html")
    path.write_bytes(bytes(msg))


def make_eml_colliding_attachments(path: Path):
    """Two distinct attachments that share one filename."""
    msg = EmailMessage()
    msg["From"] = "gilberte@example.com"
    msg["To"] = "studio@example.com"
    msg["Subject"] = "Both riders"
    msg["Date"] = "Fri, 10 Jul 2026 10:00:00 +0200"
    msg.set_content("Two riders attached, same filename.\n")
    msg.add_attachment(b"rider one\n", maintype="text", subtype="plain",
                       filename="rider.txt")
    msg.add_attachment(b"rider two\n", maintype="text", subtype="plain",
                       filename="rider.txt")
    path.write_bytes(bytes(msg))


def make_merged_xlsx(path: Path):
    """A merged title spanning two columns — the anchor value must survive."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget"
    ws["A1"] = "Q3 forecast"
    ws.merge_cells("A1:B1")
    ws.append(["item", "amount"])
    ws.append(["strings", 42.5])
    ws2 = wb.create_sheet("Notes")
    ws2["A1"] = "Wide title"
    ws2.merge_cells("A1:B1")
    ws2["A2"] = "only column A has data"
    wb.save(path)


if __name__ == "__main__":
    out = Path("fixtures-preview")
    out.mkdir(exist_ok=True)
    make_xlsx(out / "scores.xlsx"); make_csv(out / "inventory.csv")
    make_text_pdf(out / "audit.pdf"); make_scanned_pdf(out / "scan.pdf")
    make_docx(out / "procedure.docx"); make_eml(out / "mail.eml")
    make_png(out / "photo.png")
    print(f"fixtures in {out}/")
